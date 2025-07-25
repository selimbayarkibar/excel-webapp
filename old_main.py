from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import shutil
import os

app = FastAPI()

CURRENCIES = ["TL", "$", "€", "Sterlin"]


@app.post("/process-excel/")
async def process_excel(file: UploadFile = File(...)):
    original_filename = file.filename.rsplit(".", 1)[0]
    temp_file = "input.xlsx"
    output_file = f"processed_{original_filename}.xlsx"

    with open(temp_file, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    wb = load_workbook(temp_file, data_only=True)

    for currency in CURRENCIES:
        # Match kar sheet by currency + *kar (case-insensitive)
        kar_sheet_name = next(
            (name for name in wb.sheetnames if name.lower().startswith(currency.lower()) and name.lower().endswith("kar")),
            None
        )

        if currency not in wb.sheetnames or kar_sheet_name is None:
            print(f"Skipping {currency}: sheet or kar sheet not found")
            continue

        sheet = wb[currency]
        kar_sheet = wb[kar_sheet_name]
        start_row = 11

        # Clear existing data below row 10
        for row in kar_sheet.iter_rows(min_row=start_row, max_row=kar_sheet.max_row):
            for cell in row:
                cell.value = None

        kar_rows = []
        for row in sheet.iter_rows(min_row=2):
            date = row[0].value
            gelir_gider = str(row[1].value).strip().lower() if row[1].value else ""
            category = row[2].value
            explanation = row[3].value
            amount = row[4].value

            if gelir_gider not in ["gelir", "gider"] or amount is None:
                continue

            kar_rows.append((date, category, explanation, amount))

        # Write to Kar sheet
        for i, (date, cat, exp, amt) in enumerate(kar_rows):
            row_idx = start_row + i
            kar_sheet[f"A{row_idx}"] = date
            kar_sheet[f"B{row_idx}"] = cat
            kar_sheet[f"C{row_idx}"] = exp
            kar_sheet[f"D{row_idx}"] = amt

        # Insert SUM formula
        total_row = start_row + len(kar_rows) + 1
        kar_sheet[f"C{total_row}"] = "Toplam:"
        kar_sheet[f"D{total_row}"] = f"=SUM(D{start_row}:D{start_row + len(kar_rows) - 1})"

    wb.save(output_file)
    os.remove(temp_file)

    return FileResponse(output_file, filename=output_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
