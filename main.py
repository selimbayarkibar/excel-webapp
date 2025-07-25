from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import shutil
from openpyxl import load_workbook
import tempfile
import os

app = FastAPI()

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/upload/")
async def upload_files(
    source_file: UploadFile = File(...),
    destination_file: UploadFile = File(...),
    quarter: str = Form(...)
):
    # Save uploaded files temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_src:
        shutil.copyfileobj(source_file.file, temp_src)
        temp_src_path = temp_src.name

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_dest:
        shutil.copyfileobj(destination_file.file, temp_dest)
        temp_dest_path = temp_dest.name

    # Load both workbooks
    src_wb = load_workbook(temp_src_path, data_only=True)
    dest_wb = load_workbook(temp_dest_path)

    # Target sheet based on form selection
    target_sheet = f"{quarter} Comparative"

    # Currency mapping to target column
    currency_to_column = {
        "TL": "C",
        "$": "F",
        "€": "I",
    }

    # For TL, $, € only (not Sterlin)
    for currency in ["TL", "$", "€"]:
        if currency not in src_wb.sheetnames:
            continue

        src_ws = src_wb[currency]
        category_totals = {}

        # Start processing after the first 'Gelir' or 'Gider'
        start_reading = False
        for row in src_ws.iter_rows(min_row=1):
            values = [cell.value for cell in row]
            if len(values) < 5:
                continue

            t_type = values[1]
            if t_type not in ("Gelir", "Gider"):
                continue

            if not start_reading:
                start_reading = True  # We've hit the data region

            if t_type == "Gider":
                category = str(values[2]).strip()
                amount_cell = row[4]
                amount = amount_cell.value

                # Convert "(123,45)" string to -123.45 if needed
                if isinstance(amount, str) and amount.startswith("(") and amount.endswith(")"):
                    amount = float(amount.strip("()").replace(".", "").replace(",", "."))
                    amount = -abs(amount)
                elif isinstance(amount, str):
                    amount = float(amount.replace(".", "").replace(",", "."))
                elif isinstance(amount, (int, float)):
                    amount = float(amount)
                else:
                    continue

                if category in category_totals:
                    category_totals[category] += amount
                else:
                    category_totals[category] = amount

        # Write to destination sheet
        dest_ws = dest_wb[target_sheet]
        col_letter = currency_to_column[currency]

        for row in dest_ws.iter_rows(min_row=1, max_col=1):  # Column A only
            cell = row[0]
            if cell.value and str(cell.value).strip() in category_totals:
                row_idx = cell.row
                col_idx = f"{col_letter}{row_idx}"
                dest_ws[col_idx] = category_totals[str(cell.value).strip()]

    # Save and return file
    output_path = os.path.join(tempfile.gettempdir(), "processed_budget.xlsx")
    dest_wb.save(output_path)
    return FileResponse(output_path, filename="processed_budget.xlsx")
